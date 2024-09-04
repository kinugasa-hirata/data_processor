from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename
import os
import tempfile
from openpyxl import load_workbook

app = Flask(__name__)

# Use a temporary directory for uploads
UPLOAD_FOLDER = tempfile.mkdtemp()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Placeholder for the extracted data
all_extracted_data = {}
selected_data = []
selected_indices = [103, 99, 177, 158, "", "", 114, 182, 193, 79, 148, 174]

def extract_numerical_data(file_path):
    numerical_data = []
    with open(file_path, 'r') as file:
        for line in file:
            values = line.split(';')
            for value in values:
                try:
                    numeric_value = round(abs(float(value)), 2)
                    numerical_data.append(numeric_value)
                except ValueError:
                    pass
    return numerical_data

def select_data():
    global selected_data
    selected_data = []
    for file_title, data in all_extracted_data.items():
        selected_values = []
        for index in selected_indices:
            if index == "":
                selected_values.append("")
            elif index <= len(data['data']):
                selected_values.append(round(abs(data['data'][index - 1]), 2))
            else:
                selected_values.append("")
        selected_data.append({'file_title': os.path.splitext(os.path.basename(file_title))[0], 'selected_values': selected_values})
    return selected_data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})
    
    files = request.files.getlist('file')
    
    for file in files:
        if file.filename == '':
            return jsonify({'error': 'No selected file'})
        
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            numerical_data = extract_numerical_data(file_path)
            if numerical_data:
                all_extracted_data[file_path] = {"data": numerical_data}
            
            # Remove the temporary file after processing
            os.remove(file_path)
    
    selected_data = select_data()
    return jsonify({'message': 'Files processed successfully', 'data': selected_data})

@app.route('/export', methods=['POST'])
def export_to_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'})
    
    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            current_row = 17  # Start from the 17th row
            for item in selected_data:
                current_row += 1
                ws.cell(row=current_row, column=1).value = item['file_title']
                for index, value in enumerate(item['selected_values'], start=1):
                    ws.cell(row=current_row, column=index + 1).value = value
            
            # Save to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            
            # Remove the original uploaded file
            os.remove(file_path)
            
            return send_file(tmp_path, as_attachment=True, download_name='exported_data.xlsx')
        except Exception as e:
            return jsonify({'error': f'An error occurred during export: {str(e)}'})
        finally:
            # Ensure temporary files are cleaned up
            if os.path.exists(file_path):
                os.remove(file_path)
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.remove(tmp_path)

if __name__ == '__main__':
    app.run(debug=True)
